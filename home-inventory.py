# import ui package
from tkinter import *

#import excel package
from openpyxl import *

# import database package
import sqlite3

db_name = "./home-inventory/development.db"
wb1_name = "./home-inventory/export.xlsx"
wb2_name = "./home-inventory/query.xlsx"

# connect to database with name
db_con = sqlite3.connect(db_name)

# create cursor for database
db_cur = db_con.cursor()

# create initial response
db_res = db_cur.execute("SELECT name FROM sqlite_master")
db_res.fetchone()

# initialize excel sheet
wb1 = Workbook()
wb1s1 = wb1.active

wb2 = Workbook()
wb2s1 = wb2.active


def db_create():
    # execute command
    # create new main table
    db_cur.execute("CREATE TABLE Main(ID, Class, Manufacturer,\
                                      ModelNumber, SerialNumber,\
                                      DateAcquired, Description, Status, \
                                      Location, Assembly)")

def db_add_test():
    value = generate_id()
    data1 = str(int(generate_id())-99)
    data = [value, data1, data1, data1, data1,\
            data1, data1, data1, data1, data1]
    db_cur.execute("INSERT INTO Main VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data)
    db_con.commit()

def db_write():
    db_cur.execute("INSERT INTO Main VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", dataBuffer)
    db_con.commit()

def db_checkout():
    db_cur.execute("UPDATE Main SET Location = ?, Assembly = ? WHERE ID = ?", [dataBuffer[8], dataBuffer[9], dataBuffer[0]])
    db_con.commit()

def db_checkin():
    db_cur.execute("UPDATE Main SET Location = ?, Assembly = ? WHERE ID = ?", ['Storage', 'N/A', dataBuffer[0]])
    db_con.commit()

def db_modify():
    db_cur.execute("UPDATE Main SET Class = ?, Manufacturer = ?, 'ModelNumber' = ?, \
                                    'SerialNumber' = ?, 'DateAcquired' = ?, Description = ?, \
                                    Status = ?, Location = ?, Assembly = ? WHERE ID = ?", \
                                    [dataBuffer[1], dataBuffer[2], dataBuffer[3], \
                                     dataBuffer[4], dataBuffer[5], dataBuffer[6], \
                                     dataBuffer[7], dataBuffer[8], dataBuffer[9], dataBuffer[0]])
    db_con.commit()

def db_read(rowID):
    for db_row in db_cur.execute("SELECT * FROM Main WHERE ID = '" + str(rowID) + "'"):
        return(db_row)

def db_read_table():
    for db_row in db_cur.execute("SELECT * FROM Main"):
        print(db_row)

def db_delete(rowID):
    db_cur.execute("DELETE FROM Main WHERE ID = '" + str(rowID) + "'")
    db_con.commit()

def generate_id():
    id_array = []
    for db_row in db_cur.execute("SELECT ID FROM Main"):
        id_array.append(int(str(db_row).replace("'","").replace("(","").replace(")","").replace(',',"")))

    if(len(id_array) == 0):
        next_id = 100
    else:
        next_id = max(id_array) + 1
    return(str(next_id))

def db_query():
    wb2s1.delete_rows(1, wb2s1.max_row)
    
    query = query_txt.get()
    query_txt.delete(0,END)

    # print(query)
    for db_row in db_cur.execute(query):
        parsed = str(db_row).replace("'","").replace("(","").replace(")","").split(',')
        wb2s1.append(parsed)
        print(db_row)

    wb2.save(wb2_name)

# ui initalization
root = Tk()
root.title('Inventory System')

# data structures
input_text = []
dataBuffer = []

# Function to initialize 11 UI text boxes
def add_text():
    for i in range(10):
        textbox = Entry(root, width = 30, font=('Arial', 14))
        textbox.grid(row = i, column = 1)
        input_text.append(textbox)
        dataBuffer.append('-1')

# initalize text boxes
add_text()

# Immediately close the program
def exit_app():
    # close ui
    root.destroy()
    # database close
    db_con.close()

# Put data from data buffer into text boxes
def insert_data():
    # loop populate textboxes
    for i in range(len(input_text)):
        input_text[i].insert(END, dataBuffer[i])

    status.configure(text = 'Data Loaded in')

# Put data from text boxes into data buffer
def get_data():
    # loop retrieve input text
    for i in range(len(input_text)):
        dataBuffer[i] = input_text[i].get()

    status.configure(text = 'Data Loaded into buffer')

# clear text boxes        
def clear_data():
    # update status message
    for i in range(len(input_text)):
        input_text[i].delete(0,END)

    index_txt.delete(0,END)
    status.configure(text = 'Data Cleared')

def write_data():
    # load buffer
    get_data()

    # generate next ID
    dataBuffer[0] = generate_id()
    
    # write buffer to database
    db_write()

    # clear data
    clear_data()

    # update status
    status.configure(text = 'Writing complete')

def read_data():
    # check read id
    read_id = index_txt.get()

    # query database
    test = db_read(read_id)
    
    # copy query into data buffer
    for i in range(len(input_text)):
        dataBuffer[i] = test[i]

    # display data from buffer
    clear_data()
    insert_data()
    
    # set status
    status.configure(text = 'Data read')

def delete_data():
    # check delete id
    delete_id = index_txt.get()

    # query database
    test = db_read(delete_id)
    
    # copy query into data buffer
    for i in range(len(input_text)):
        dataBuffer[i] = test[i]

    # display data from buffer
    clear_data()
    insert_data()

    # query database
    db_delete(delete_id)
    
    # set status
    status.configure(text = 'Data deleted')

def export_data():
    wb1s1.delete_rows(1, wb1s1.max_row)
    
    for db_row in db_cur.execute("SELECT * FROM Main"):
        parsed = str(db_row).replace("'","").replace("(","").replace(")","").split(',')
        wb1s1.append(parsed)

    wb1.save(wb1_name)

def check_out():
    # load buffer
    get_data()

    # modify database using buffer
    db_checkout()

    # clear data
    clear_data()

    # update status
    status.configure(text = 'Item checked out')

def check_in():
    # load buffer
    get_data()

    # modify database using buffer
    db_checkin()

    # clear data
    clear_data()

    # update status
    status.configure(text = 'Item checked in')

def modify_data():
    # load buffer
    get_data()

    # modify database using buffer
    db_modify()

    # clear data
    clear_data()

    # update status
    status.configure(text = 'Data modified')

# ID
lbl1 = Label(root, text = "ID", justify = LEFT)
lbl1.grid(row = 0, column = 0)

# Class
lbl2 = Label(root, text = "Class", justify = LEFT)
lbl2.grid(row = 1, column = 0)
 
# Manufacturer
lbl4 = Label(root, text = "Manufacturer", justify = LEFT)
lbl4.grid(row = 2, column = 0)

# Model Number
lbl5 = Label(root, text = "Part Number", justify = LEFT)
lbl5.grid(row = 3, column = 0)

# Serial Number
lbl6 = Label(root, text = "Serial Number", justify = LEFT)
lbl6.grid(row = 4, column = 0)

# Date Acquired
lbl7 = Label(root, text = "Date Acquired", justify = LEFT)
lbl7.grid(row = 5, column = 0)

# Description
lbl8 = Label(root, text = "Description", justify = LEFT)
lbl8.grid(row = 6, column = 0)

# Status
lbl9 = Label(root, text = "Status", justify = LEFT)
lbl9.grid(row = 7, column = 0)

# Location
lbl10 = Label(root, text = "Location", justify = LEFT)
lbl10.grid(row = 8, column = 0)

# Assembly
lbl11 = Label(root, text = "Assembly", justify = LEFT)
lbl11.grid(row = 9, column = 0)

# Status Message
status = Label(root, text = "", justify = LEFT)
status.grid(row = 10, column = 1)

# index text box
index_txt = Entry(root, width = 30, font=('Arial', 14))
index_txt.grid(column = 1, row = 11)

# Index
index_lbl = Label(root, text = "Index", justify = LEFT)
index_lbl.grid(row = 11, column = 0)

# Enter Data Button
enter_bttn = Button(root, text = "Enter", fg = "black", bg = "white", width = 10, command = write_data)
enter_bttn.grid(column = 2, row = 0)

# Clear Button
clear_bttn = Button(root, text = "Clear", fg = "black", bg = "white", width = 10, command = clear_data)
clear_bttn.grid(column = 2, row = 1)

# Read Button
read_bttn = Button(root, text = "Read", fg = "black", bg = "white", width = 10, command = read_data)
read_bttn.grid(column = 2, row = 2)

# Delete Button
delete_bttn = Button(root, text = "Delete", fg = "black", bg = "white", width = 10, command = delete_data)
delete_bttn.grid(column = 2, row = 3)

# Modify Button
modify_bttn = Button(root, text = "Modify", fg = "black", bg = "white", width = 10, command = modify_data)
modify_bttn.grid(column = 2, row = 4)

#Check in Button
checkin_bttn = Button(root, text = "Check in", fg = "black", bg = "white", width = 10, command = check_in)
checkin_bttn.grid(column = 2, row = 7)

# Check out Button
checkout_bttn = Button(root, text = "Check out", fg = "black", bg = "white", width = 10, command = check_out)
checkout_bttn.grid(column = 2, row = 8)

# Export Button
export_bttn = Button(root, text = "Export", fg = "black", bg = "white", width = 10, command = export_data)
export_bttn.grid(column = 2, row = 9)

# Exit Button
exit_bttn = Button(root, text = "Exit", fg = "black", bg = "white", width = 10, command = exit_app)
exit_bttn.grid(column = 2, row = 11)


# Query Label
query_lbl = Label(root, text = "SQL", justify = LEFT)
query_lbl.grid(row = 13, column = 0)

# Query Text Box
query_txt = Entry(root, width = 30, font=('Arial', 14))
query_txt.grid(row = 13, column = 1)

# Query Button
query_bttn = Button(root, text = "Query", fg = "black", bg = "white", width = 10, command = db_query)
query_bttn.grid(row = 13, column = 2)


# db_create()
# db_add_test()
# db_read_table()

root.mainloop()